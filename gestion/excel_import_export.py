import re
import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from gestion.models import Persona, Alumno, Carrera, Materia, PlanEstudio, Comision, Cursada

def normalizar_dni(dni_raw):
    """Limpia puntos, guiones y espacios de un DNI."""
    if not dni_raw:
        return ""
    return re.sub(r'\D', '', str(dni_raw)).strip()

def normalizar_cuil(cuil_raw):
    """Limpia puntos, guiones y espacios de un CUIL."""
    if not cuil_raw:
        return ""
    return re.sub(r'\D', '', str(cuil_raw)).strip()

def normalizar_nombre_propio(texto):
    """Formatea nombres y apellidos con mayúscula inicial en cada palabra."""
    if not texto:
        return ""
    
    # Reemplazo de tildes invertidas
    reemplazos_tildes = {
        'à': 'á', 'è': 'é', 'ì': 'í', 'ò': 'ó', 'ù': 'ú',
        'À': 'Á', 'È': 'É', 'Ì': 'Í', 'Ò': 'Ó', 'Ù': 'Ú'
    }
    limpio = str(texto).strip()
    for origen, destino in reemplazos_tildes.items():
        limpio = limpio.replace(origen, destino)
        
    palabras = limpio.split()
    palabras_formateadas = []
    for p in palabras:
        if len(p) <= 2 and p.lower() in ['de', 'del', 'la', 'las', 'los', 'y', 'e', 'da', 'do']:
            palabras_formateadas.append(p.lower())
        else:
            palabras_formateadas.append(p.capitalize())
    
    if palabras_formateadas:
        palabras_formateadas[0] = palabras_formateadas[0].capitalize()
    return " ".join(palabras_formateadas)

def parsear_fecha_flexible(fecha_raw):
    """Parsea una fecha en diversos formatos a datetime.date."""
    if not fecha_raw:
        return None
    if isinstance(fecha_raw, datetime.date):
        return fecha_raw
    if isinstance(fecha_raw, datetime.datetime):
        return fecha_raw.date()

    val_str = str(fecha_raw).strip()
    formatos = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
        '%d/%m/%y', '%d-%m-%y', '%d.%m.%Y'
    ]
    for fmt in formatos:
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None

def normalizar_genero(genero_raw):
    """Normaliza la identidad de género a las opciones válidas: M, F, X, N."""
    if not genero_raw:
        return 'N'
    g_str = str(genero_raw).strip().lower()
    if g_str in ['m', 'masculino', 'varón', 'varon', 'hombre', 'male']:
        return 'M'
    elif g_str in ['f', 'femenino', 'mujer', 'female']:
        return 'F'
    elif g_str in ['x', 'no binario', 'no-binario', 'nobinario', 'otro', 'otra', 'diversidad']:
        return 'X'
    return 'N'

def normalizar_mail(mail_raw):
    """Valida y limpia el correo electrónico."""
    if not mail_raw:
        return None
    m_str = str(mail_raw).strip().lower()
    if not m_str or m_str in ['none', 'null', '-', '--', 's/d', 'sin dato', 'no tiene']:
        return None
    if re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', m_str):
        return m_str
    return None

def generar_plantilla_alumnos_excel():
    """
    Genera un archivo Excel oficial (.xlsx) diseñado específicamente para directivos
    con instrucciones claras, validaciones y formato intuitivo para la carga de estudiantes.
    """
    wb = openpyxl.Workbook()
    
    # HOJA 1: CARGA DE ESTUDIANTES
    ws = wb.active
    ws.title = "Carga de Estudiantes"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    sample_font = Font(name="Arial", size=9, italic=True, color="475569")
    data_font = Font(name="Arial", size=10, color="0F172A")
    inst_font = Font(name="Arial", size=9, bold=True, color="1E3A8A")
    inst_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Título institucional
    ws.merge_cells("A1:M1")
    t_cell = ws["A1"]
    t_cell.value = "INSTITUTO SUPERIOR DE FORMACIÓN TÉCNICA N° 188 — PLANTILLA OFICIAL DE CARGA DE ESTUDIANTES"
    t_cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila de Instrucciones
    ws.merge_cells("A2:M2")
    i_cell = ws["A2"]
    i_cell.value = "📌 INSTRUCCIONES: Complete una fila por estudiante. DNI, Apellido y Nombre son obligatorios. Los demás campos son opcionales (si no cuenta con el dato, déjelo vacío)."
    i_cell.font = inst_font
    i_cell.fill = inst_fill
    i_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24

    # Encabezados de Columnas
    headers = [
        ("DNI (*)", "Obligatorio (ej: 45039996 o 45.039.996)"),
        ("Apellido (*)", "Obligatorio (ej: Gómez, Pérez)"),
        ("Nombre (*)", "Obligatorio (ej: Juan Carlos)"),
        ("Carrera", "Nombre o Código de la Tecnicatura (ej: Energía, Higiene, Logística)"),
        ("Año de Cursada", "1, 2 o 3 (Año lectivo en la carrera)"),
        ("CUIL", "Opcional (ej: 20450399966)"),
        ("Fecha de Nacimiento", "Opcional (Formato DD/MM/AAAA ej: 15/04/2001)"),
        ("Identidad de Género", "Opcional (Masculino, Femenino, Otro, Prefiere no decir)"),
        ("Nacionalidad", "Opcional (por defecto Argentina)"),
        ("Localidad", "Opcional (ej: General Rodríguez, Moreno, Luján)"),
        ("Domicilio", "Opcional (ej: Av. España 1234)"),
        ("Teléfono / Celular", "Opcional (ej: 1123456789)"),
        ("Correo Electrónico", "Opcional (ej: estudiante@gmail.com — si no tiene déjelo vacío)")
    ]

    ws.row_dimensions[3].height = 28
    for col_idx, (h_title, comment_txt) in enumerate(headers, start=1):
        cell = ws.cell(3, col_idx)
        cell.value = h_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Filas de Ejemplo
    ejemplos = [
        ("45039996", "Abeldaño Aquino", "Gonzalo Daniel", "Tecnicatura Superior en Energía", "1", "20450399966", "13/01/2000", "Masculino", "Argentina", "Moreno", "Angel Gallardo 6160", "1130858866", "gonzalodaniel2107@gmail.com"),
        ("39110038", "Alegre", "Aldana Micaela", "Tecnicatura Superior en Higiene", "2", "27391100384", "18/09/1995", "Femenino", "Argentina", "General Rodríguez", "México Manzana 3 Casa 41", "1155443322", ""),
        ("34412371", "Acosta", "María Noemí", "Tecnicatura Superior en Logística", "1", "", "", "Prefiere no decir", "Argentina", "General Rodríguez", "", "", "")
    ]

    for row_idx, fila in enumerate(ejemplos, start=4):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(fila, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = sample_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # HOJA 2: REFERENCIA DE CARRERAS
    ws_ref = wb.create_sheet(title="Carreras Disponibles")
    ws_ref.views.sheetView[0].showGridLines = True
    
    ws_ref.merge_cells("A1:C1")
    t_ref = ws_ref["A1"]
    t_ref.value = "LISTADO DE CARRERAS OFICIALES — ISFT N° 188"
    t_ref.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    t_ref.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    t_ref.alignment = Alignment(horizontal="center", vertical="center")
    ws_ref.row_dimensions[1].height = 26

    ref_headers = ["Código de Carrera", "Nombre de la Tecnicatura", "Resolución Vigente"]
    for c_idx, h_text in enumerate(ref_headers, start=1):
        cell = ws_ref.cell(2, c_idx, h_text)
        cell.font = header_font
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_ref.row_dimensions[2].height = 22

    carreras_db = Carrera.objects.all().order_by('nombre_carrera')
    for r_idx, c in enumerate(carreras_db, start=3):
        ws_ref.cell(r_idx, 1, c.codigo_carrera).border = thin_border
        ws_ref.cell(r_idx, 2, c.nombre_carrera).border = thin_border
        ws_ref.cell(r_idx, 3, c.resolucion_vigente or "S/R").border = thin_border

    # Anchos de columna automáticos
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    for col in ws_ref.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ref.column_dimensions[col_letter].width = max(max_len + 4, 20)

    # Validaciones de Lista
    dv_genero = DataValidation(type="list", formula1='"Masculino,Femenino,Otro,Prefiere no decir"', allow_blank=True)
    ws.add_data_validation(dv_genero)
    dv_genero.add("H4:H1000")

    dv_anio = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    ws.add_data_validation(dv_anio)
    dv_anio.add("E4:E1000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def procesar_importacion_alumnos_excel(file_obj):
    """
    Procesa un archivo Excel subido por el usuario para registrar/actualizar estudiantes.
    Es flexible ante datos faltantes (CUIL, correos, localidad, teléfono) y 
    reporta con precisión filas procesadas, creadas, actualizadas y observaciones.
    """
    summary = {
        'success': False,
        'creados': 0,
        'actualizados': 0,
        'filas_procesadas': 0,
        'errores': [],
        'advertencias': [],
        'mensaje': ''
    }

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        summary['mensaje'] = f"El archivo proporcionado no es un archivo Excel (.xlsx) válido: {str(e)}"
        return summary

    ws = None
    for sheet_name in ["Carga de Estudiantes", "Estudiantes", "Carga de Alumnos", "Alumnos", "Hoja 1", "Sheet1"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    if not ws:
        ws = wb.active

    if not ws or ws.max_row < 2:
        summary['mensaje'] = "La hoja de cálculo no contiene filas suficientes de datos para procesar."
        return summary

    # Detectar fila de encabezados buscando 'DNI' o 'DOCUMENTO'
    header_row = None
    col_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
        if any("DNI" in v or "DOCUMENTO" in v for v in row_vals):
            header_row = r
            for c_idx, val in enumerate(row_vals, 1):
                if "DNI" in val or "DOCUMENTO" in val: col_map["dni"] = c_idx
                elif "APELLIDO" in val: col_map["apellido"] = c_idx
                elif "NOMBRE" in val and "CARRERA" not in val: col_map["nombre"] = c_idx
                elif "CARRERA" in val or "TECNICATURA" in val: col_map["carrera"] = c_idx
                elif "AÑO" in val or "ANIO" in val or "CURSADA" in val: col_map["anio"] = c_idx
                elif "CUIL" in val: col_map["cuil"] = c_idx
                elif "FECHA" in val or "NACIMIENTO" in val: col_map["fecha_nac"] = c_idx
                elif "GENERO" in val or "GÉNERO" in val or "IDENTIDAD" in val: col_map["genero"] = c_idx
                elif "NACIONALIDAD" in val: col_map["nacionalidad"] = c_idx
                elif "LOCALIDAD" in val or "CIUDAD" in val: col_map["localidad"] = c_idx
                elif "DOMICILIO" in val or "DIRECCION" in val or "DIRECCIÓN" in val: col_map["domicilio"] = c_idx
                elif "TELEFONO" in val or "TELÉFONO" in val or "CELULAR" in val: col_map["telefono"] = c_idx
                elif "MAIL" in val or "CORREO" in val or "EMAIL" in val: col_map["mail"] = c_idx
            break

    if not header_row or "dni" not in col_map or "apellido" not in col_map or "nombre" not in col_map:
        summary['mensaje'] = "No se pudieron identificar las columnas obligatorias (DNI, Apellido, Nombre) en el archivo Excel. Utilice la plantilla oficial."
        return summary

    carreras_cache = list(Carrera.objects.all())

    # Procesar filas de datos
    for r in range(header_row + 1, ws.max_row + 1):
        dni_raw = ws.cell(r, col_map["dni"]).value
        apellido_raw = ws.cell(r, col_map["apellido"]).value
        nombre_raw = ws.cell(r, col_map["nombre"]).value

        # Si toda la fila obligatoria está vacía, continuar
        if not dni_raw and not apellido_raw and not nombre_raw:
            continue

        dni_clean = normalizar_dni(dni_raw)
        if not dni_clean:
            summary['errores'].append(f"Fila {r}: DNI inválido o ausente ('{dni_raw}').")
            continue

        apellido_clean = normalizar_nombre_propio(apellido_raw)
        nombre_clean = normalizar_nombre_propio(nombre_raw)

        if not apellido_clean or not nombre_clean:
            summary['errores'].append(f"Fila {r} (DNI {dni_clean}): Falta completar el Apellido o Nombre del estudiante.")
            continue

        # Campos Opcionales Flexibles
        cuil_raw = ws.cell(r, col_map["cuil"]).value if "cuil" in col_map else None
        cuil_clean = normalizar_cuil(cuil_raw) or None

        fecha_nac_raw = ws.cell(r, col_map["fecha_nac"]).value if "fecha_nac" in col_map else None
        fecha_nac_clean = parsear_fecha_flexible(fecha_nac_raw)

        genero_raw = ws.cell(r, col_map["genero"]).value if "genero" in col_map else None
        genero_clean = normalizar_genero(genero_raw)

        nacionalidad_raw = ws.cell(r, col_map["nacionalidad"]).value if "nacionalidad" in col_map else None
        nacionalidad_clean = str(nacionalidad_raw).strip() if nacionalidad_raw else "Argentina"

        localidad_raw = ws.cell(r, col_map["localidad"]).value if "localidad" in col_map else None
        localidad_clean = str(localidad_raw).strip() if localidad_raw else None

        domicilio_raw = ws.cell(r, col_map["domicilio"]).value if "domicilio" in col_map else None
        domicilio_clean = str(domicilio_raw).strip() if domicilio_raw else None

        telefono_raw = ws.cell(r, col_map["telefono"]).value if "telefono" in col_map else None
        telefono_clean = str(telefono_raw).strip() if telefono_raw else None

        mail_raw = ws.cell(r, col_map["mail"]).value if "mail" in col_map else None
        mail_clean = normalizar_mail(mail_raw)

        carrera_raw = ws.cell(r, col_map["carrera"]).value if "carrera" in col_map else None
        anio_raw = ws.cell(r, col_map["anio"]).value if "anio" in col_map else None

        # Crear o actualizar Persona
        persona_obj, created = Persona.objects.update_or_create(
            dni=dni_clean,
            defaults={
                'apellido': apellido_clean,
                'nombre': nombre_clean,
                'cuil': cuil_clean,
                'fecha_nacimiento': fecha_nac_clean,
                'identidad': genero_clean,
                'nacionalidad': nacionalidad_clean,
                'localidad': localidad_clean,
                'domicilio': domicilio_clean,
                'telefono': telefono_clean,
                'mail': mail_clean
            }
        )

        if created:
            summary['creados'] += 1
        else:
            summary['actualizados'] += 1

        # Crear o actualizar Alumno
        legajo_gen = f"LEG-{dni_clean}"
        alumno_obj, _ = Alumno.objects.get_or_create(
            persona=persona_obj,
            defaults={'legajo': legajo_gen}
        )

        # Asignar Cursada/Carrera si se especificó
        if carrera_raw:
            carrera_str = str(carrera_raw).strip().lower()
            carrera_match = None
            for c in carreras_cache:
                if (c.codigo_carrera.lower() in carrera_str or 
                    c.nombre_carrera.lower() in carrera_str or 
                    carrera_str in c.nombre_carrera.lower() or 
                    carrera_str in c.codigo_carrera.lower()):
                    carrera_match = c
                    break

            if carrera_match:
                try:
                    anio_num = int(str(anio_raw).replace('°', '').replace('º', '').strip()) if anio_raw else 1
                    if anio_num not in [1, 2, 3]:
                        anio_num = 1
                except (ValueError, TypeError):
                    anio_num = 1
                
                planes = PlanEstudio.objects.filter(carrera=carrera_match, anio_carrera=anio_num)
                if not planes.exists():
                    planes = PlanEstudio.objects.filter(carrera=carrera_match)
                
                for plan in planes:
                    com = Comision.objects.filter(plan_estudio=plan).first()
                    if not com:
                        com, _ = Comision.objects.get_or_create(
                            codigo_comision=f"COM_{plan.id_plan}_2026",
                            defaults={'plan_estudio': plan, 'anio_lectivo': 2026}
                        )
                    Cursada.objects.get_or_create(
                        comision=com,
                        alumno=alumno_obj,
                        defaults={'situacion_final': 'Regular', 'porcentaje_asistencia': 80.0}
                    )
            else:
                summary['advertencias'].append(f"Fila {r} (DNI {dni_clean}): La carrera '{carrera_raw}' no coincidió con ninguna carrera registrada, por lo que el estudiante se registró sin cursada automática.")

        summary['filas_procesadas'] += 1

    summary['success'] = True
    summary['mensaje'] = f"Importación finalizada con éxito: {summary['creados']} estudiantes nuevos registrados y {summary['actualizados']} estudiantes actualizados."
    return summary
