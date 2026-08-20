import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Carrera, Alumno, Cursada, PlanEstudio, Materia, Evaluacion

def generar_libro_matriz_excel(carrera_obj):
    """
    Genera el archivo Excel oficial del Libro Matriz para una Carrera específica
    siguiendo la estructura estándar del ISFT N° 188.
    Retorna un objeto BytesIO con el archivo .xlsx generado.
    """
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Estilos compartidos
    font_title = Font(name='Arial', size=14, bold=True, color='1E3A8A')
    font_subtitle = Font(name='Arial', size=11, bold=True, color='475569')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_data = Font(name='Arial', size=9)
    font_data_bold = Font(name='Arial', size=9, bold=True)
    
    fill_header_navy = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    fill_header_blue = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    fill_header_indigo = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_success = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border_side = Side(border_style='thin', color='CBD5E1')
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Obtener alumnos inscritos en la carrera
    alumnos_carrera = Alumno.objects.filter(
        cursadas__comision__plan_estudio__carrera=carrera_obj
    ).select_related('persona').distinct().order_by('persona__apellido', 'persona__nombre')

    # Planes y materias organizadas por año (1°, 2°, 3°)
    planes_carrera = PlanEstudio.objects.filter(carrera=carrera_obj).select_related('materia').order_by('anio_carrera', 'id_plan')
    materias_por_anio = {1: [], 2: [], 3: []}
    for p in planes_carrera:
        an = p.anio_carrera if p.anio_carrera in [1, 2, 3] else 1
        materias_por_anio[an].append(p)

    # -------------------------------------------------------------
    # 1. HOJA: LIBRO MATRIZ (Resumen General)
    # -------------------------------------------------------------
    ws_matriz = wb.create_sheet(title='LIBRO MATRIZ')
    ws_matriz.views.sheetView[0].showGridLines = True
    
    # Encabezado Institucional
    ws_matriz.merge_cells('A1:G1')
    ws_matriz['A1'] = f"INSTITUTO SUPERIOR DE FORMACIÓN TÉCNICA N° 188 — LIBRO MATRIZ"
    ws_matriz['A1'].font = font_title
    ws_matriz['A1'].alignment = align_center

    ws_matriz.merge_cells('A2:G2')
    ws_matriz['A2'] = f"CARRERA: {carrera_obj.nombre_carrera} (Resolución: {carrera_obj.resolucion_vigente or 'S/R'})"
    ws_matriz['A2'].font = font_subtitle
    ws_matriz['A2'].alignment = align_center

    headers_matriz = ['Alumno', 'DNI', 'Libro', 'Folio / Legajo', 'Cohorte', 'Egresado', 'Plan de Estudio']
    for col_num, h_text in enumerate(headers_matriz, 1):
        cell = ws_matriz.cell(row=4, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_cell

    row_curr = 5
    for idx, al in enumerate(alumnos_carrera, 1):
        p = al.persona
        # Determinar si egreso (ej: si tiene todas las materias promocionadas o final)
        curs_al = Cursada.objects.filter(alumno=al, comision__plan_estudio__carrera=carrera_obj)
        total_cursadas = curs_al.count()
        promos_or_finals = curs_al.filter(situacion_final__in=['Promocionado', 'Final']).count()
        es_egresado = "SI" if (total_cursadas > 0 and total_cursadas == promos_or_finals and total_cursadas >= len(planes_carrera) and len(planes_carrera) > 0) else "NO"
        
        libro_num = (idx // 100) + 1
        folio_num = (idx % 100) or 100

        vals = [
            f"{p.apellido}, {p.nombre}",
            p.dni,
            libro_num,
            al.legajo or f"F-{folio_num}",
            "2024",
            es_egresado,
            carrera_obj.resolucion_vigente or carrera_obj.codigo_carrera
        ]
        
        bg_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        if es_egresado == "SI":
            bg_fill = fill_success

        for col_num, val in enumerate(vals, 1):
            cell = ws_matriz.cell(row=row_curr, column=col_num, value=val)
            cell.font = font_data
            if bg_fill.fill_type: cell.fill = bg_fill
            cell.border = border_cell
            cell.alignment = align_center if col_num in [2, 3, 4, 5, 6] else align_left
        row_curr += 1

    # Ajustar anchos de columnas
    for col in ws_matriz.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_matriz.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------
    # 2. HOJAS POR ANO: 1° Año, 2° Año, 3° Año
    # -------------------------------------------------------------
    for anio in [1, 2, 3]:
        sheet_title = f"{anio}° Año"
        ws_anio = wb.create_sheet(title=sheet_title)
        ws_anio.views.sheetView[0].showGridLines = True
        
        planes_anio = materias_por_anio.get(anio, [])
        total_cols = 3 + len(planes_anio)

        ws_anio.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 4))
        ws_anio.cell(row=1, column=1, value=f"PLANILLA DE CALIFICACIONES — {anio}° AÑO ({carrera_obj.nombre_carrera})").font = font_title
        ws_anio.cell(row=1, column=1).alignment = align_center

        # Headers
        headers_anio = ['Estudiante', 'DNI', 'Folio']
        for p in planes_anio:
            headers_anio.append(p.materia.nombre_materia)

        for col_num, h_text in enumerate(headers_anio, 1):
            cell = ws_anio.cell(row=3, column=col_num, value=h_text)
            cell.font = font_header
            cell.fill = fill_header_blue if anio == 1 else (fill_header_indigo if anio == 2 else fill_header_navy)
            cell.alignment = align_center
            cell.border = border_cell

        row_curr = 4
        for idx, al in enumerate(alumnos_carrera, 1):
            p = al.persona
            folio_num = al.legajo or f"{idx}"
            row_data = [f"{p.apellido}, {p.nombre}", p.dni, folio_num]

            # Buscar notas de cada materia en este año
            for plan_it in planes_anio:
                curs = Cursada.objects.filter(alumno=al, comision__plan_estudio=plan_it).first()
                if curs:
                    # Obtener ultima evaluacion o situacion
                    ev = curs.evaluaciones.filter(nota__isnull=False).order_by('-fecha').first()
                    if ev and ev.nota is not None:
                        row_data.append(float(ev.nota))
                    elif curs.situacion_final:
                        row_data.append(curs.situacion_final)
                    else:
                        row_data.append("")
                else:
                    row_data.append("")

            bg_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
            for col_num, val in enumerate(row_data, 1):
                cell = ws_anio.cell(row=row_curr, column=col_num, value=val)
                cell.font = font_data
                if bg_fill.fill_type: cell.fill = bg_fill
                cell.border = border_cell
                cell.alignment = align_center if col_num > 1 else align_left
            row_curr += 1

        for col in ws_anio.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_anio.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # -------------------------------------------------------------
    # 3. HOJA: EGRESADOS
    # -------------------------------------------------------------
    ws_egr = wb.create_sheet(title='Egresados')
    ws_egr.views.sheetView[0].showGridLines = True
    
    ws_egr.merge_cells('A1:F1')
    ws_egr.cell(row=1, column=1, value=f"PLANILLA DE EGRESADOS — {carrera_obj.nombre_carrera}").font = font_title
    ws_egr.cell(row=1, column=1).alignment = align_center

    headers_egr = ['Estudiante', 'DNI', 'Libro', 'Folio', 'Egreso', 'Año de Egreso']
    for col_num, h_text in enumerate(headers_egr, 1):
        cell = ws_egr.cell(row=3, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_cell

    row_curr = 4
    for idx, al in enumerate(alumnos_carrera, 1):
        p = al.persona
        curs_al = Cursada.objects.filter(alumno=al, comision__plan_estudio__carrera=carrera_obj)
        total_cursadas = curs_al.count()
        promos_or_finals = curs_al.filter(situacion_final__in=['Promocionado', 'Final']).count()
        es_egr = "SI" if (total_cursadas > 0 and total_cursadas == promos_or_finals and total_cursadas >= len(planes_carrera) and len(planes_carrera) > 0) else "NO"
        
        anio_egreso = "2026" if es_egr == "SI" else ""
        libro_num = (idx // 100) + 1
        folio_num = al.legajo or f"{idx}"

        vals = [f"{p.apellido}, {p.nombre}", p.dni, libro_num, folio_num, es_egr, anio_egreso]
        bg_fill = fill_success if es_egr == "SI" else (fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None))

        for col_num, val in enumerate(vals, 1):
            cell = ws_egr.cell(row=row_curr, column=col_num, value=val)
            cell.font = font_data
            if bg_fill.fill_type: cell.fill = bg_fill
            cell.border = border_cell
            cell.alignment = align_center if col_num > 1 else align_left
        row_curr += 1

    for col in ws_egr.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_egr.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
