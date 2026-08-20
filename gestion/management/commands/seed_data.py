import os
import openpyxl
import datetime
import re
from django.core.management.base import BaseCommand
from django.core.management import call_command
from django.db import transaction
from gestion.models import Persona, Alumno, Docente, Carrera, Materia, PlanEstudio, Comision, ComisionDocente, Cursada, Evaluacion

class Command(BaseCommand):
    help = 'Poblar la base de datos a partir de los archivos Excel normalizados en las carpetas Recursos y excels'

    GENDER_MAP = {
        'masculino': 'M',
        'm': 'M',
        'varon': 'M',
        'varón': 'M',
        'hombre': 'M',
        'femenino': 'F',
        'f': 'F',
        'mujer': 'F',
        'otro': 'O',
        'no binario': 'O',
        'nobinario': 'O',
        'o': 'O',
        'prefiere no decir': 'N',
        'prefiero no decir': 'N',
        'prefiero no contestar': 'N',
        'no contesta': 'N',
        'n': 'N',
        '': 'N',
        'none': 'N'
    }

    OFFICIAL_CARRERAS = [
        {"codigo": "ENERGIA-794", "nombre": "Técnico Superior en Energía con Orientación Industrial", "res": "Res. 794/01"},
        {"codigo": "ESTERILIZACION-530", "nombre": "Tecnicatura Superior en Tecnología en Salud con especialidad en Esterilización", "res": "Res. 530/09"},
        {"codigo": "HIGIENE-320", "nombre": "Tecnicatura Superior en Higiene y Seguridad en el Trabajo", "res": "Res. 320/13"},
        {"codigo": "HIGIENE-6183", "nombre": "Tecnicatura Superior en Higiene y Seguridad en el Trabajo", "res": "Res. 6183/25"},
        {"codigo": "LOGISTICA-1557", "nombre": "Tecnicatura Superior en Logística", "res": "Res. 1557/08"},
        {"codigo": "LOGISTICA-5312", "nombre": "Tecnicatura Superior en Logística", "res": "Res. 5312/24"},
        {"codigo": "LABORATORIO-205", "nombre": "Técnico Superior en Laboratorio de Análisis Clínicos", "res": "Res. 205/18"},
        {"codigo": "LABORATORIO-6182", "nombre": "Técnico Superior en Laboratorio de Análisis Clínicos", "res": "Res. 6182/25"},
        {"codigo": "DEPORTIVAS-2786", "nombre": "Tecnicatura Superior en Prácticas Deportivas", "res": "Res. 2786/20"},
        {"codigo": "ENFERMERIA-854", "nombre": "Tecnicatura Superior en Enfermería", "res": "Res. 854/16"},
    ]

    def fix_accents(self, text):
        """
        Corrige tildes invertidas (ej: è -> é, à -> á) y artefactos de codificación.
        """
        if not text:
            return text
        replacements = {
            'à': 'á', 'è': 'é', 'ì': 'í', 'ò': 'ó', 'ù': 'ú',
            'À': 'Á', 'È': 'É', 'Ì': 'Í', 'Ò': 'Ó', 'Ù': 'Ú',
            'â': 'á', 'ê': 'é', 'î': 'í', 'ô': 'ó', 'û': 'ú',
            'Â': 'Á', 'Ê': 'É', 'Î': 'Í', 'Ô': 'Ó', 'Û': 'Ú',
            'ã': 'a', 'õ': 'o',
            'Ã': 'A', 'Õ': 'O',
        }
        text_str = str(text)
        for bad, good in replacements.items():
            text_str = text_str.replace(bad, good)
        return text_str

    def clean_title_case(self, text):
        """
        Estandariza nombres y apellidos en Title Case respetando conectores en minúscula (de, del, la, y).
        Ejemplo: 'ABELDAÑO AQUINO' -> 'Abeldaño Aquino', 'DE BANDI' -> 'De Bandi'
        """
        if not text or str(text).strip() in ['', 'None', '-', 'nan', 'S/D', 'S/N', 'S/A']:
            return ""
        text = self.fix_accents(str(text).strip())
        lower_connectors = {'de', 'del', 'la', 'las', 'el', 'los', 'y', 'e', 'da', 'di', 'do', 'das', 'dos', 'van', 'von', 'd'}
        words = text.split()
        result = []
        for i, w in enumerate(words):
            if '-' in w:
                subwords = [sw.capitalize() for sw in w.split('-')]
                w_cap = '-'.join(subwords)
            elif "'" in w:
                subwords = [sw.capitalize() for sw in w.split("'")]
                w_cap = "'".join(subwords)
            else:
                w_lower = w.lower()
                if i > 0 and w_lower in lower_connectors:
                    w_cap = w_lower
                else:
                    w_cap = w.capitalize()
            result.append(w_cap)
        return " ".join(result)

    def clean_gender(self, val):
        if not val:
            return 'N'
        v = str(val).strip().lower()
        return self.GENDER_MAP.get(v, 'N')

    def parse_date(self, val):
        if not val or str(val).strip() in ['', 'None', '-', '--']:
            return None
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        s = str(val).strip().split(' ')[0]
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def clean_phone(self, val):
        if not val or str(val).strip() in ['', 'None', '-']:
            return None
        s = str(val).strip()
        if s.endswith('.0'):
            s = s[:-2]
        return s

    def clean_cuil(self, val):
        if not val or str(val).strip() in ['', 'None', '-']:
            return None
        s = str(val).strip()
        if s.endswith('.0'):
            s = s[:-2]
        return s.replace('-', '').replace('.', '').replace(' ', '')

    @transaction.atomic
    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS("Iniciando purga y sembrado masivo de datos normalizados..."))
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        recursos_dir = os.path.join(base_dir, 'Recursos')
        excels_dir = os.path.join(base_dir, 'excels')

        # 1. PURGA DE DATOS VIEJOS
        self.stdout.write("Purgando registros anteriores...")
        Evaluacion.objects.all().delete()
        Cursada.objects.all().delete()
        ComisionDocente.objects.all().delete()
        Comision.objects.all().delete()
        PlanEstudio.objects.all().delete()
        Materia.objects.all().delete()
        Docente.objects.all().delete()
        Alumno.objects.all().delete()
        Persona.objects.all().delete()
        Carrera.objects.all().delete()

        # 2. CARGA DE CARRERAS OFICIALES Y MATERIAS
        self.stdout.write("Procesando catálogo oficial de Carreras y Materias...")
        for c_dict in self.OFFICIAL_CARRERAS:
            Carrera.objects.update_or_create(
                codigo_carrera=c_dict["codigo"],
                defaults={
                    "nombre_carrera": self.fix_accents(c_dict["nombre"]),
                    "resolucion_vigente": c_dict["res"]
                }
            )

        carreras_file = os.path.join(recursos_dir, 'Carreras_y_Materias_ISFT188.xlsx')
        if os.path.exists(carreras_file):
            wb_carr = openpyxl.load_workbook(carreras_file, data_only=True)
            if 'Materias_Planes_Estudio' in wb_carr.sheetnames:
                ws_mat = wb_carr['Materias_Planes_Estudio']
                for r in range(4, ws_mat.max_row + 1):
                    nom_carrera_sheet = self.fix_accents(str(ws_mat.cell(r, 1).value or '').strip())
                    resolucion_sheet = str(ws_mat.cell(r, 2).value or '').strip()
                    cod_mat_raw = str(ws_mat.cell(r, 3).value or '').strip()
                    anio_c = ws_mat.cell(r, 4).value
                    nom_materia = self.fix_accents(str(ws_mat.cell(r, 5).value or '').strip())
                    carga_anual = ws_mat.cell(r, 6).value
                    carga_sem = ws_mat.cell(r, 7).value
                    correlativas = ws_mat.cell(r, 8).value

                    if not cod_mat_raw or not nom_materia or nom_materia == 'None':
                        continue

                    # Vincular con Carrera
                    carrera_obj = None
                    if nom_carrera_sheet:
                        for c_cand in Carrera.objects.all():
                            if c_cand.nombre_carrera.lower() in nom_carrera_sheet.lower() or nom_carrera_sheet.lower() in c_cand.nombre_carrera.lower():
                                carrera_obj = c_cand
                                break
                    if not carrera_obj:
                        carrera_obj = Carrera.objects.filter(codigo_carrera="LOGISTICA-5312").first() or Carrera.objects.first()

                    # Codigo unico por materia y carrera
                    codigo_materia_full = f"{carrera_obj.codigo_carrera}_{cod_mat_raw}"
                    nom_materia_clean = self.fix_accents(nom_materia)
                    materia_obj, _ = Materia.objects.update_or_create(
                        codigo_materia=codigo_materia_full,
                        defaults={'nombre_materia': nom_materia_clean}
                    )

                    # Plan de Estudio
                    try:
                        anio_int = int(str(anio_c).replace('°', '').replace('º', '').strip())
                    except (ValueError, TypeError):
                        anio_int = 1

                    try:
                        c_anual = int(float(str(carga_anual).strip())) if carga_anual else None
                    except (ValueError, TypeError):
                        c_anual = None

                    try:
                        c_semanal = int(float(str(carga_sem).strip())) if carga_sem else None
                    except (ValueError, TypeError):
                        c_semanal = None

                    plan_obj, _ = PlanEstudio.objects.get_or_create(
                        carrera=carrera_obj,
                        materia=materia_obj,
                        defaults={
                            'anio_carrera': anio_int,
                            'carga_horaria_anual': c_anual,
                            'carga_horaria_semanal': c_semanal,
                            'correlatividades': str(correlativas or '').strip()
                        }
                    )

                    com_obj, _ = Comision.objects.get_or_create(
                        codigo_comision=f"COM_{plan_obj.id_plan}_2026",
                        defaults={
                            'plan_estudio': plan_obj,
                            'anio_lectivo': 2026,
                            'cuatrimestre': 'Anual',
                            'turno': 'Vespertino',
                            'division': 'A'
                        }
                    )

        # 3. CARGA MASIVA DE ALUMNOS DESDE excels/
        carrera_file_mapping = {
            "Deportivas": "DEPORTIVAS-2786",
            "Energia": "ENERGIA-794",
            "Enfemeria": "ENFERMERIA-854",
            "Esterilizacion": "ESTERILIZACION-530",
            "Laboratorio": "LABORATORIO-205",
            "Logistica": "LOGISTICA-5312",
            "Seguridad": "HIGIENE-320",
        }

        if not os.path.exists(excels_dir):
            self.stdout.write(self.style.ERROR("Directorio 'excels' no encontrado."))
            return

        self.stdout.write("\nProcesando planillas maestras de estudiantes...")
        total_alumnos_procesados = 0
        total_cursadas_creadas = 0

        for fname in sorted(os.listdir(excels_dir)):
            if fname.startswith("~$") or not fname.endswith(".xlsx") or fname == "LIBRO MATRIZ DE ENERGÍA.xlsx":
                continue

            carrera_tag = None
            for k in carrera_file_mapping:
                if k.lower() in fname.lower():
                    carrera_tag = k
                    break

            cod_carrera = carrera_file_mapping.get(carrera_tag, "ENERGIA-794")
            carrera_obj = Carrera.objects.filter(codigo_carrera=cod_carrera).first() or Carrera.objects.first()

            fpath = os.path.join(excels_dir, fname)
            self.stdout.write(f"  -> Ingestando {fname} ({carrera_obj.nombre_carrera})...")
            wb = openpyxl.load_workbook(fpath, data_only=True)

            for sname in wb.sheetnames:
                ws = wb[sname]
                
                anio_cursada = 1
                if '2' in sname:
                    anio_cursada = 2
                elif '3' in sname:
                    anio_cursada = 3

                header_row = None
                col_map = {}
                for r in range(1, min(10, ws.max_row + 1)):
                    row_vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                    if "DNI" in [v.upper() for v in row_vals]:
                        header_row = r
                        for c_idx, val in enumerate(row_vals, 1):
                            v_u = val.upper()
                            if "CORREO" in v_u or "EMAIL" in v_u or "MAIL" in v_u: col_map["mail"] = c_idx
                            elif "APELLIDO" in v_u: col_map["apellido"] = c_idx
                            elif "NOMBRE" in v_u: col_map["nombre"] = c_idx
                            elif "DNI" in v_u: col_map["dni"] = c_idx
                            elif "CUIL" in v_u: col_map["cuil"] = c_idx
                            elif "GENERO" in v_u or "IDENTIDAD" in v_u: col_map["genero"] = c_idx
                            elif "NACIONALIDAD" in v_u: col_map["nacionalidad"] = c_idx
                            elif "FECHA" in v_u or "NACIMIENTO" in v_u: col_map["fecha_nac"] = c_idx
                            elif "CELULAR" in v_u or "TELEFONO" in v_u: col_map["telefono"] = c_idx
                            elif "DIRECCI" in v_u or "DOMICILIO" in v_u: col_map["domicilio"] = c_idx
                            elif "LOCALIDAD" in v_u: col_map["localidad"] = c_idx
                        break

                if not header_row:
                    continue

                for r in range(header_row + 1, ws.max_row + 1):
                    dni_val = ws.cell(r, col_map.get("dni", 4)).value if "dni" in col_map else None
                    if not dni_val or str(dni_val).strip() in ["", "None", "TOTAL"]:
                        continue
                    try:
                        dni_clean = str(int(float(str(dni_val).strip())))
                    except (ValueError, TypeError):
                        continue

                    ap_val = str(ws.cell(r, col_map.get("apellido", 2)).value or "").strip()
                    nom_val = str(ws.cell(r, col_map.get("nombre", 3)).value or "").strip()
                    mail_val = str(ws.cell(r, col_map.get("mail", 1)).value or "").strip()
                    cuil_raw = ws.cell(r, col_map.get("cuil", 5)).value
                    gen_raw = ws.cell(r, col_map.get("genero", 6)).value
                    nac_raw = str(ws.cell(r, col_map.get("nacionalidad", 7)).value or "").strip()
                    fnac_raw = ws.cell(r, col_map.get("fecha_nac", 8)).value
                    tel_raw = ws.cell(r, col_map.get("telefono", 10)).value
                    dom_raw = str(ws.cell(r, col_map.get("domicilio", 11)).value or "").strip()
                    loc_raw = str(ws.cell(r, col_map.get("localidad", 12)).value or "").strip()

                    cuil_clean = self.clean_cuil(cuil_raw)
                    gen_sigla = self.clean_gender(gen_raw)
                    fnac_clean = self.parse_date(fnac_raw)
                    tel_clean = self.clean_phone(tel_raw)
                    nacionalidad_clean = self.clean_title_case(nac_raw) if nac_raw and nac_raw != 'None' else 'Argentina'
                    domicilio_clean = self.fix_accents(dom_raw) if dom_raw and dom_raw != 'None' else None
                    localidad_clean = self.clean_title_case(loc_raw) if loc_raw and loc_raw != 'None' else None
                    
                    # CORREOS: Solo guardar si es un correo real del estudiante, nunca inventar
                    mail_clean = mail_val if mail_val and '@' in mail_val and not mail_val.endswith('@isft188.edu.ar') and mail_val != 'None' else None

                    # Nombres y Apellidos en Title Case estandarizado
                    nom_clean = self.clean_title_case(nom_val)
                    ap_clean = self.clean_title_case(ap_val)

                    persona_obj, created = Persona.objects.get_or_create(
                        dni=dni_clean,
                        defaults={
                            'cuil': cuil_clean,
                            'nombre': nom_clean,
                            'apellido': ap_clean,
                            'domicilio': domicilio_clean,
                            'localidad': localidad_clean,
                            'telefono': tel_clean,
                            'mail': mail_clean,
                            'nacionalidad': nacionalidad_clean,
                            'fecha_nacimiento': fnac_clean,
                            'identidad': gen_sigla
                        }
                    )

                    if not created:
                        if not persona_obj.cuil and cuil_clean: persona_obj.cuil = cuil_clean
                        if not persona_obj.fecha_nacimiento and fnac_clean: persona_obj.fecha_nacimiento = fnac_clean
                        if persona_obj.identidad == 'N' and gen_sigla != 'N': persona_obj.identidad = gen_sigla
                        if not persona_obj.localidad and localidad_clean: persona_obj.localidad = localidad_clean
                        if not persona_obj.domicilio and domicilio_clean: persona_obj.domicilio = domicilio_clean
                        if not persona_obj.telefono and tel_clean: persona_obj.telefono = tel_clean
                        if nom_clean and not persona_obj.nombre: persona_obj.nombre = nom_clean
                        if ap_clean and not persona_obj.apellido: persona_obj.apellido = ap_clean
                        if mail_clean and not persona_obj.mail: persona_obj.mail = mail_clean
                        persona_obj.save()

                    alumno_obj, _ = Alumno.objects.get_or_create(
                        persona=persona_obj,
                        defaults={'legajo': f"LEG-{dni_clean}"}
                    )
                    total_alumnos_procesados += 1

                    planes_anio = PlanEstudio.objects.filter(carrera=carrera_obj, anio_carrera=anio_cursada)
                    if not planes_anio.exists():
                        planes_anio = PlanEstudio.objects.filter(carrera=carrera_obj)

                    for plan_it in planes_anio:
                        com_it = Comision.objects.filter(plan_estudio=plan_it).first()
                        if not com_it:
                            com_it, _ = Comision.objects.get_or_create(
                                codigo_comision=f"COM_{plan_it.id_plan}_2026",
                                defaults={'plan_estudio': plan_it, 'anio_lectivo': 2026}
                            )
                        curs_obj, curs_created = Cursada.objects.get_or_create(
                            comision=com_it,
                            alumno=alumno_obj,
                            defaults={
                                'porcentaje_asistencia': 80.0,
                                'situacion_final': 'Regular'
                            }
                        )
                        if curs_created:
                            total_cursadas_creadas += 1

        # 4. INGESTA DEL LIBRO MATRIZ DE ENERGIA (Calificaciones, Cursadas Historicas y Egresados)
        libro_matriz_path = os.path.join(excels_dir, 'LIBRO MATRIZ DE ENERGÍA.xlsx')
        if os.path.exists(libro_matriz_path):
            self.stdout.write("\nProcesando LIBRO MATRIZ DE ENERGÍA (Evaluaciones y Calificaciones)...")
            wb_lm = openpyxl.load_workbook(libro_matriz_path, data_only=True)
            carrera_energia = Carrera.objects.filter(codigo_carrera="ENERGIA-794").first()
            total_evals_cargadas = 0

            # Mapear egresados
            egresados_dict = {}
            if 'Egresados' in wb_lm.sheetnames:
                ws_eg = wb_lm['Egresados']
                for r in range(4, ws_eg.max_row + 1):
                    dni_eg_val = ws_eg.cell(r, 2).value
                    eg_val = str(ws_eg.cell(r, 5).value or "").strip().upper()
                    if dni_eg_val and eg_val == 'SI':
                        try:
                            dni_eg_cl = str(int(float(str(dni_eg_val).strip())))
                            egresados_dict[dni_eg_cl] = True
                        except (ValueError, TypeError):
                            pass

            for sname in ['1º Año', '2º Año', '3º Año']:
                if sname not in wb_lm.sheetnames:
                    continue
                ws_lm = wb_lm[sname]
                anio_lm = 1 if '1' in sname else (2 if '2' in sname else 3)
                
                # Mapear columnas de materias
                col_mat_map = {}
                for c in range(4, ws_lm.max_column + 1):
                    header_val = str(ws_lm.cell(3, c).value or "").strip()
                    if header_val and not header_val.startswith('Corr.') and header_val != 'Promedio':
                        clean_mat_name = header_val.split('-')[-1].strip().lower()
                        clean_mat_name = self.fix_accents(clean_mat_name)
                        mat_obj = None
                        for m_cand in Materia.objects.filter(codigo_materia__startswith="ENERGIA"):
                            m_cand_nom = self.fix_accents(m_cand.nombre_materia.lower())
                            if clean_mat_name in m_cand_nom or m_cand_nom in clean_mat_name:
                                mat_obj = m_cand
                                break
                        if mat_obj:
                            col_mat_map[c] = mat_obj

                # Leer filas de notas y estudiantes
                for r in range(4, ws_lm.max_row + 1):
                    dni_v = ws_lm.cell(r, 2).value
                    nom_completo = str(ws_lm.cell(r, 1).value or "").strip()
                    if not dni_v:
                        continue
                    try:
                        dni_cl = str(int(float(str(dni_v).strip())))
                    except (ValueError, TypeError):
                        continue

                    # Obtener o crear Persona y Alumno si es historico
                    persona_lm = Persona.objects.filter(dni=dni_cl).first()
                    if not persona_lm:
                        nom_cleaned = self.clean_title_case(nom_completo)
                        parts = nom_cleaned.split(' ')
                        if len(parts) >= 4:
                            ap = f"{parts[0]} {parts[1]}"
                            nm = " ".join(parts[2:])
                        elif len(parts) == 3:
                            ap = parts[0]
                            nm = f"{parts[1]} {parts[2]}"
                        elif len(parts) == 2:
                            ap = parts[0]
                            nm = parts[1]
                        else:
                            ap = parts[0] if parts else "S/A"
                            nm = ""
                        
                        persona_lm = Persona.objects.create(
                            dni=dni_cl,
                            nombre=nm,
                            apellido=ap,
                            identidad='N',
                            mail=None
                        )
                    else:
                        # Estandarizar nombre si ya existia
                        persona_lm.nombre = self.clean_title_case(persona_lm.nombre)
                        persona_lm.apellido = self.clean_title_case(persona_lm.apellido)
                        if persona_lm.mail and persona_lm.mail.endswith('@isft188.edu.ar'):
                            persona_lm.mail = None
                        persona_lm.save()

                    alumno_lm, _ = Alumno.objects.get_or_create(
                        persona=persona_lm,
                        defaults={'legajo': f"LEG-{dni_cl}"}
                    )

                    for col_idx, mat_obj in col_mat_map.items():
                        nota_val = ws_lm.cell(r, col_idx).value
                        if nota_val is not None and str(nota_val).strip() != '':
                            try:
                                nota_num = float(nota_val)
                            except (ValueError, TypeError):
                                nota_num = None

                            plan_obj = PlanEstudio.objects.filter(carrera=carrera_energia, materia=mat_obj).first()
                            if not plan_obj:
                                plan_obj, _ = PlanEstudio.objects.get_or_create(
                                    carrera=carrera_energia,
                                    materia=mat_obj,
                                    defaults={'anio_carrera': anio_lm}
                                )

                            com_obj = Comision.objects.filter(plan_estudio=plan_obj).first()
                            if not com_obj:
                                com_obj, _ = Comision.objects.get_or_create(
                                    codigo_comision=f"COM_{plan_obj.id_plan}_2026",
                                    defaults={'plan_estudio': plan_obj, 'anio_lectivo': 2026}
                                )

                            situacion = "Promocionado" if (nota_num is not None and nota_num >= 7) else ("Regular" if (nota_num is not None and nota_num >= 4) else "Regular")
                            cursada_lm, _ = Cursada.objects.update_or_create(
                                comision=com_obj,
                                alumno=alumno_lm,
                                defaults={
                                    'porcentaje_asistencia': 85.0 if situacion == "Promocionado" else 75.0,
                                    'situacion_final': situacion
                                }
                            )

                            if nota_num is not None and nota_num > 0:
                                Evaluacion.objects.update_or_create(
                                    cursada=cursada_lm,
                                    instancia="Nota Final",
                                    defaults={
                                        'nota': nota_num,
                                        'fecha': datetime.date(2026, 7, 15)
                                    }
                                )
                                total_evals_cargadas += 1

        # 5. GENERACION AUTOMATICA DE FIXTURE CANONICA (initial_data.json)
        fixtures_dir = os.path.join(base_dir, 'gestion', 'fixtures')
        os.makedirs(fixtures_dir, exist_ok=True)
        fixture_path = os.path.join(fixtures_dir, 'initial_data.json')
        
        self.stdout.write(f"\nExportando fixture institucional canónica: {fixture_path}...")
        from io import StringIO
        out = StringIO()
        call_command('dumpdata', 'gestion', indent=2, stdout=out)
        with open(fixture_path, 'w', encoding='utf-8') as f:
            f.write(out.getvalue())

        total_personas = Persona.objects.count()
        total_alumnos = Alumno.objects.count()
        total_carreras = Carrera.objects.count()
        total_materias = Materia.objects.count()
        total_cursadas = Cursada.objects.count()
        total_evals = Evaluacion.objects.count()

        self.stdout.write(self.style.SUCCESS("\n======================================================="))
        self.stdout.write(self.style.SUCCESS("¡Sembrado completado con éxito bajo Estándares de Calidad!"))
        self.stdout.write(self.style.SUCCESS("======================================================="))
        self.stdout.write(f"• Total Personas en BD: {total_personas}")
        self.stdout.write(f"• Total Alumnos en BD: {total_alumnos}")
        self.stdout.write(f"• Total Carreras: {total_carreras}")
        self.stdout.write(f"• Total Materias: {total_materias}")
        self.stdout.write(f"• Total Cursadas / Inscripciones: {total_cursadas}")
        self.stdout.write(f"• Total Evaluaciones / Calificaciones: {total_evals}")
        self.stdout.write(self.style.SUCCESS("======================================================="))
